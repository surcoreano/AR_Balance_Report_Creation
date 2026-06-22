import os
import re
import calendar
import threading
import platform
import pandas as pd
import numpy as np
import tkinter as tk

from datetime import datetime
from tkinter import filedialog, messagebox, ttk

# ─────────────────────────────
# Theme & Typography (UI 디자인 설정)
# ─────────────────────────────
if platform.system() == "Darwin":
    FN = "Arial"
    FM = "Menlo"
else:
    FN = "Arial"
    FM = "Consolas"

APP_BG = "#f4f6f9"
CARD_BG = "#ffffff"
PRIMARY = "#1a73e8"
PRIMARY_DARK = "#1557b0"
TEXT = "#202124"
SUBTEXT = "#5f6368"
BORDER = "#dadce0"
LOG_BG = "#f8f9fa"

TEXTS = {
    "EN": {
        "title": "AR Balance Report Generator",
        "header": "AR Balance Report Generator",
        "subtitle": "Format Excel/TSV files and generate PDF & Excel automatically.",
        "language": "Language",
        "excel_file": "Data file (Excel/TSV)",
        "output_folder": "Output folder",
        "due_date": "Payment Due Date",
        "due_date_hint": "Example: 260625 (YYMMDD)",
        "browse": "Browse",
        "start": "Start process",
        "ready": "Select a data file and output folder.",
        "log": "Result Log",
        "excel_error": "Please select a valid Excel or TSV file.",
        "folder_error": "Please select a valid Output folder.",
        "date_error": "Please enter the date in YYMMDD format.",
        "confirm_title": "Confirmation",
        "confirm_msg": "Do you want to start generating the report?",
        "cancelled": "Process cancelled.",
        "error": "Error",
        "done": "Completed",
        "reading_excel": "Reading and analyzing data file...",
        "processing": "Formatting and filtering data...",
        "saving": "Applying styles and saving Excel file...",
        "completed_status": "Completed successfully. Files saved.",
        "completed_msg": "Excel & PDF report generation is complete.\n\nSaved location:\n{}"
    },
    "ES": {
        "title": "Generador de Reportes AR Balance",
        "header": "Generador de Reportes AR Balance",
        "subtitle": "Formatee archivos de datos y genere Excel y PDFs automáticamente.",
        "language": "Idioma",
        "excel_file": "Archivo de datos (Excel/TSV)",
        "output_folder": "Carpeta de destino",
        "due_date": "Fecha de Vencimiento",
        "due_date_hint": "Ejemplo: 260625 (YYMMDD)",
        "browse": "Buscar",
        "start": "Iniciar proceso",
        "ready": "Seleccione un archivo de datos y una carpeta.",
        "log": "Registro de resultados",
        "excel_error": "Seleccione un archivo Excel o TSV válido.",
        "folder_error": "Seleccione una carpeta de destino válida.",
        "date_error": "Ingrese la fecha en formato YYMMDD.",
        "confirm_title": "Confirmación",
        "confirm_msg": "¿Desea iniciar la generación del reporte?",
        "cancelled": "Proceso cancelado.",
        "error": "Error",
        "done": "Completado",
        "reading_excel": "Leyendo y analizando archivo de datos...",
        "processing": "Formateando y filtrando datos...",
        "saving": "Generando PDFs y archivo Excel...",
        "completed_status": "Completado con éxito. Archivos guardados.",
        "completed_msg": "La generación de Excel y PDFs ha finalizado.\n\nUbicación:\n{}"
    },
    "KR": {
        "title": "AR Balance 리포트 생성기",
        "header": "AR Balance 리포트 생성기",
        "subtitle": "데이터를 포맷팅하고 Excel 및 PDF 요약본을 자동으로 생성합니다.",
        "language": "언어",
        "excel_file": "원본 파일 (Excel/TSV)",
        "output_folder": "결과물 저장 폴더",
        "due_date": "Payment Due Date",
        "due_date_hint": "예시: 260625 (YYMMDD 형식)",
        "browse": "찾아보기",
        "start": "실행 시작",
        "ready": "원본 파일과 저장할 폴더를 선택해주세요.",
        "log": "진행 로그",
        "excel_error": "올바른 Excel 또는 TSV 파일을 선택해주세요.",
        "folder_error": "올바른 저장 폴더를 선택해주세요.",
        "date_error": "날짜를 정확히 YYMMDD 형식으로 입력해주세요.",
        "confirm_title": "확인",
        "confirm_msg": "리포트 생성을 시작하시겠습니까?",
        "cancelled": "작업이 취소되었습니다.",
        "error": "오류",
        "done": "완료",
        "reading_excel": "데이터를 읽고 분석하는 중...",
        "processing": "데이터를 포맷팅하고 필터링하는 중...",
        "saving": "Excel 서식 지정 및 PDF 생성 중...",
        "completed_status": "성공적으로 완료되었습니다.",
        "completed_msg": "Excel 및 개별 PDF 생성이 완료되었습니다.\n\n저장 위치:\n{}"
    }
}

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.lang = tk.StringVar(value="ES")
        
        self.excel_file = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.due_date = tk.StringVar()

        self.geometry("750x700")
        self.configure(bg=APP_BG)

        self._setup_style()
        self._build_ui()
        self._apply_language()

    def t(self, key):
        return TEXTS[self.lang.get()][key]

    def _setup_style(self):
        style = ttk.Style(self)
        try: style.theme_use("clam")
        except tk.TclError: pass

        style.configure("Google.Horizontal.TProgressbar", troughcolor="#e8f0fe", background=PRIMARY, thickness=6)
        style.configure("Google.TCombobox", fieldbackground=APP_BG, background=APP_BG, padding=4)

    def _build_ui(self):
        top_frame = tk.Frame(self, bg=APP_BG, padx=30, pady=20)
        top_frame.pack(fill="x")

        title_inner = tk.Frame(top_frame, bg=APP_BG)
        title_inner.pack(side="left")
        self.header_lbl = tk.Label(title_inner, font=(FN, 20, "bold"), bg=APP_BG, fg=TEXT)
        self.header_lbl.pack(anchor="w")
        self.subtitle_lbl = tk.Label(title_inner, font=(FN, 10), bg=APP_BG, fg=SUBTEXT)
        self.subtitle_lbl.pack(anchor="w", pady=(5, 0))

        lang_inner = tk.Frame(top_frame, bg=APP_BG)
        lang_inner.pack(side="right", anchor="ne")
        self.language_lbl = tk.Label(lang_inner, font=(FN, 9), bg=APP_BG, fg=SUBTEXT)
        self.language_lbl.pack(anchor="w")
        self.language_combo = ttk.Combobox(lang_inner, textvariable=self.lang, values=["EN", "ES", "KR"], state="readonly", width=6, style="Google.TCombobox")
        self.language_combo.pack(pady=(2, 0))
        self.language_combo.bind("<<ComboboxSelected>>", lambda e: self._apply_language())

        main_frame = tk.Frame(self, bg=CARD_BG, padx=30, pady=30, highlightbackground=BORDER, highlightthickness=1)
        main_frame.pack(fill="x", padx=20)

        self.excel_lbl = tk.Label(main_frame, font=(FN, 9, "bold"), bg=CARD_BG, fg=TEXT)
        self.excel_lbl.pack(anchor="w")
        excel_row = tk.Frame(main_frame, bg=CARD_BG)
        excel_row.pack(fill="x", pady=(5, 15))
        self.excel_entry = tk.Entry(excel_row, textvariable=self.excel_file, font=(FN, 10), relief="solid", bd=1)
        self.excel_entry.pack(side="left", fill="x", expand=True, ipady=6)
        self.excel_btn = tk.Button(excel_row, font=(FN, 9, "bold"), bg=CARD_BG, fg=PRIMARY, width=10, relief="solid", bd=1, command=self._browse_excel)
        self.excel_btn.pack(side="left", padx=(10, 0), ipady=4)

        self.folder_lbl = tk.Label(main_frame, font=(FN, 9, "bold"), bg=CARD_BG, fg=TEXT)
        self.folder_lbl.pack(anchor="w")
        folder_row = tk.Frame(main_frame, bg=CARD_BG)
        folder_row.pack(fill="x", pady=(5, 15))
        self.folder_entry = tk.Entry(folder_row, textvariable=self.output_folder, font=(FN, 10), relief="solid", bd=1)
        self.folder_entry.pack(side="left", fill="x", expand=True, ipady=6)
        self.folder_btn = tk.Button(folder_row, font=(FN, 9, "bold"), bg=CARD_BG, fg=PRIMARY, width=10, relief="solid", bd=1, command=self._browse_folder)
        self.folder_btn.pack(side="left", padx=(10, 0), ipady=4)

        self.date_lbl = tk.Label(main_frame, font=(FN, 9, "bold"), bg=CARD_BG, fg=TEXT)
        self.date_lbl.pack(anchor="w")
        date_row = tk.Frame(main_frame, bg=CARD_BG)
        date_row.pack(fill="x", pady=(5, 2))
        self.date_entry = tk.Entry(date_row, textvariable=self.due_date, font=(FN, 10), relief="solid", bd=1, width=20)
        self.date_entry.pack(side="left", ipady=6)
        self.date_hint_lbl = tk.Label(date_row, font=(FN, 8), bg=CARD_BG, fg=SUBTEXT)
        self.date_hint_lbl.pack(side="left", padx=(10, 0))

        log_frame = tk.Frame(self, bg=CARD_BG, padx=20, pady=20, highlightbackground=BORDER, highlightthickness=1)
        log_frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.log_lbl = tk.Label(log_frame, font=(FN, 9, "bold"), bg=CARD_BG, fg=TEXT)
        self.log_lbl.pack(anchor="w", pady=(0, 5))
        
        self.log = tk.Text(log_frame, height=6, font=(FM, 9), bg=LOG_BG, fg=TEXT, relief="solid", bd=1, padx=10, pady=10, state="disabled")
        self.log.pack(fill="both", expand=True)

        bottom_frame = tk.Frame(self, bg=APP_BG, padx=20, pady=10)
        bottom_frame.pack(fill="x", side="bottom")

        self.progress_var = tk.IntVar(value=0)
        status_inner = tk.Frame(bottom_frame, bg=APP_BG)
        status_inner.pack(side="left", fill="x", expand=True, padx=(0, 20))
        
        self.progress = ttk.Progressbar(status_inner, variable=self.progress_var, maximum=100, style="Google.Horizontal.TProgressbar")
        self.progress.pack(fill="x")
        self.status_lbl = tk.Label(status_inner, font=(FN, 9), bg=APP_BG, fg=PRIMARY, anchor="w")
        self.status_lbl.pack(fill="x", pady=(5, 0))

        self.run_btn = tk.Button(bottom_frame, font=(FN, 10, "bold"), bg=PRIMARY, fg="white", activebackground=PRIMARY_DARK, activeforeground="white", relief="flat", padx=20, pady=10, command=self._run)
        self.run_btn.pack(side="right")

    def _apply_language(self):
        self.title(self.t("title"))
        self.header_lbl.config(text=self.t("header"))
        self.subtitle_lbl.config(text=self.t("subtitle"))
        self.language_lbl.config(text=self.t("language"))
        self.excel_lbl.config(text=self.t("excel_file"))
        self.folder_lbl.config(text=self.t("output_folder"))
        self.date_lbl.config(text=self.t("due_date"))
        self.date_hint_lbl.config(text=self.t("due_date_hint"))
        self.excel_btn.config(text=self.t("browse"))
        self.folder_btn.config(text=self.t("browse"))
        self.run_btn.config(text=self.t("start"))
        self.status_lbl.config(text=self.t("ready"))
        self.log_lbl.config(text=self.t("log"))

    def _browse_excel(self):
        path = filedialog.askopenfilename(
            title=self.t("excel_file"), 
            filetypes=[("Data files", "*.xlsx *.xls *.tsv *.csv *.txt"), ("All files", "*.*")]
        )
        if path: self.excel_file.set(path)

    def _browse_folder(self):
        path = filedialog.askdirectory(title=self.t("output_folder"))
        if path: self.output_folder.set(path)

    def _log(self, message):
        self.log.config(state="normal")
        self.log.insert("end", message + "\n")
        self.log.see("end")
        self.log.config(state="disabled")

    def _set_status(self, pct, text):
        self.progress_var.set(pct)
        self.status_lbl.config(text=text)
        self.update_idletasks()

    def _clear_log(self):
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")

    def _run(self):
        file_path = self.excel_file.get().strip()
        output_dir = self.output_folder.get().strip()
        due_date_str = self.due_date.get().strip()

        if not file_path or not os.path.isfile(file_path):
            messagebox.showerror(self.t("error"), self.t("excel_error"))
            return
        if not output_dir or not os.path.isdir(output_dir):
            messagebox.showerror(self.t("error"), self.t("folder_error"))
            return
        
        try:
            payment_due_date = datetime.strptime(due_date_str, "%y%m%d")
        except ValueError:
            messagebox.showerror(self.t("error"), self.t("date_error"))
            return

        if not messagebox.askyesno(self.t("confirm_title"), self.t("confirm_msg")):
            self.status_lbl.config(text=self.t("cancelled"))
            return

        self.run_btn.config(state="disabled", bg="#9aa0a6")
        self.progress_var.set(0)
        self._clear_log()
        self._log(f"[{datetime.now().strftime('%H:%M:%S')}] Started processing...")

        def worker():
            try:
                self._set_status(10, self.t("reading_excel"))
                self._log(f"Reading target file: {os.path.basename(file_path)}")
                
                file_ext = os.path.splitext(file_path)[1].lower()
                is_tsv = file_ext in ['.tsv', '.txt', '.csv']
                separator = ',' if file_ext == '.csv' else '\t'
                
                def safe_read_csv(path, sep, header_row=None, rows=None):
                    encodings_to_try = ['utf-8', 'cp1252', 'iso-8859-1', 'utf-16', 'utf-16-le', 'cp949', 'euc-kr']
                    for enc in encodings_to_try:
                        try:
                            return pd.read_csv(path, sep=sep, header=header_row, nrows=rows, encoding=enc, low_memory=False, thousands=',')
                        except Exception:
                            continue
                    return pd.read_csv(path, sep=sep, header=header_row, nrows=rows, encoding='utf-8', encoding_errors='replace', low_memory=False, thousands=',')

                if is_tsv:
                    temp_df = safe_read_csv(file_path, separator, header_row=None, rows=30)
                else:
                    temp_df = pd.read_excel(file_path, header=None, nrows=30)
                
                header_idx = 0
                for i, row in temp_df.iterrows():
                    if any(pd.notna(val) and 'Invoice No' in str(val) for val in row.values):
                        header_idx = i
                        break
                
                self._log(f"Header found at row {header_idx + 1}. Loading full data...")
                
                if is_tsv:
                    original_df = safe_read_csv(file_path, separator, header_row=header_idx)
                else:
                    original_df = pd.read_excel(file_path, header=header_idx)
                    
                df = original_df.copy()

                self._set_status(30, self.t("processing"))
                
                fill_cols = ['Reference No', 'PO No', 'Comments', 'Invoice Remark']
                for col in fill_cols:
                    if col in df.columns:
                        df[col] = df[col].replace(r'^\s*$', np.nan, regex=True)

                if 'Reference No' not in df.columns:
                    df['Reference No'] = np.nan

                if 'PO No' in df.columns: df['Reference No'] = df['Reference No'].fillna(df['PO No'])
                if 'Comments' in df.columns: df['Reference No'] = df['Reference No'].fillna(df['Comments'])
                if 'Invoice Remark' in df.columns: df['Reference No'] = df['Reference No'].fillna(df['Invoice Remark'])

                df.columns = [' '.join(str(c).split()) for c in df.columns]

                columns_to_keep = [
                    'Invoice No.', 'AR Class', 'Trx Date', 'Due Date', 
                    'Original Amount (Entered Curr.)', 'Balance Total', 
                    'Bill To Code', 'Bill To Name', 'Reference No'
                ]
                existing_cols = [col for col in columns_to_keep if col in df.columns]
                df = df[existing_cols]

                rename_dict = {
                    'Invoice No.': 'Factura', 'AR Class': 'Tipología', 'Trx Date': 'Fecha emisión',
                    'Due Date': 'Fecha Vencimiento', 'Original Amount (Entered Curr.)': 'Importe',
                    'Balance Total': 'Balance', 'Bill To Code': 'Código Cliente',
                    'Bill To Name': 'Nombre Cliente', 'Reference No': 'Referencia'
                }
                df = df.rename(columns=rename_dict)

                ar_class_mapping = {'CM': 'Abono', 'DM': 'Devolución recibo', 'INV': 'Factura', 'PMT': 'Importe a cuenta'}
                if 'Tipología' in df.columns:
                    df['Tipología'] = df['Tipología'].replace(ar_class_mapping)

                # 💡 [요구사항 1 반영] 기준일의 '월말(Last day of the month)'을 계산하여 그 이후의 데이터만 필터링
                last_day = calendar.monthrange(payment_due_date.year, payment_due_date.month)[1]
                max_allowed_date = datetime(payment_due_date.year, payment_due_date.month, last_day)

                self._log(f"Filtering dates strictly up to end of month: {max_allowed_date.strftime('%Y-%m-%d')}")
                
                if 'Fecha Vencimiento' in df.columns:
                    df['Fecha Vencimiento'] = pd.to_datetime(df['Fecha Vencimiento'], errors='coerce')
                    df = df[df['Fecha Vencimiento'] <= max_allowed_date]

                sort_cols = [col for col in ['Nombre Cliente', 'Fecha Vencimiento'] if col in df.columns]
                if sort_cols:
                    df = df.sort_values(by=sort_cols, ascending=[True]*len(sort_cols))

                for date_col in ['Fecha emisión', 'Fecha Vencimiento']:
                    if date_col in df.columns:
                        df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')

                target_order = ['Código Cliente', 'Nombre Cliente', 'Tipología', 'Factura', 'Fecha emisión', 'Fecha Vencimiento', 'Importe', 'Balance', 'Referencia']
                final_order = [col for col in target_order if col in df.columns]
                df = df[final_order]

                for num_col in ['Importe', 'Balance']:
                    if num_col in df.columns:
                        if df[num_col].dtype == 'object':
                            df[num_col] = df[num_col].replace({',': ''}, regex=True)
                        df[num_col] = pd.to_numeric(df[num_col], errors='coerce').fillna(0)

                if 'Código Cliente' in df.columns and 'Nombre Cliente' in df.columns and 'Balance' in df.columns:
                    df_summary = df.groupby(['Código Cliente', 'Nombre Cliente'], as_index=False)['Balance'].sum()
                    df_summary = df_summary.sort_values(by='Nombre Cliente', ascending=True)
                else:
                    df_summary = pd.DataFrame(columns=['Código Cliente', 'Nombre Cliente', 'Balance'])

                # 💡 [요구사항 2 반영] 거래선별 PDF 자동 생성 로직
                self._set_status(60, "Generating PDF reports...")
                self._log("Creating individual PDF files...")
                
                pdf_dir = os.path.join(output_dir, "PDF files")
                os.makedirs(pdf_dir, exist_ok=True)
                
                try:
                    from reportlab.lib.pagesizes import A4, landscape
                    from reportlab.lib import colors
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    
                    styles = getSampleStyleSheet()
                    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor("#1557b0"), spaceAfter=15, alignment=0)
                    
                    pdf_cols = ['Tipología', 'Factura', 'Fecha emisión', 'Fecha Vencimiento', 'Importe', 'Balance', 'Referencia']
                    available_pdf_cols = [c for c in pdf_cols if c in df.columns]
                    
                    for (client_code, client_name), group_df in df.groupby(['Código Cliente', 'Nombre Cliente']):
                        clean_name = re.sub(r'[\\/*?:"<>|]', "", str(client_name))
                        
                        dates = pd.to_datetime(group_df['Fecha Vencimiento'], errors='coerce')
                        max_date_val = dates.max()
                        max_date_str = max_date_val.strftime('%Y-%m-%d') if pd.notna(max_date_val) else ""
                        
                        total_bal = group_df['Balance'].sum()
                        total_bal_fmt = f"({abs(total_bal):,.2f})" if total_bal < 0 else f"{total_bal:,.2f}"
                        total_bal_filename = f"{total_bal:,.2f}".replace(",", "") # 파일명에 콤마 제거
                        
                        pdf_filename = f"{client_code}_{clean_name}_{max_date_str}_{total_bal_filename}.pdf"
                        pdf_filepath = os.path.join(pdf_dir, pdf_filename)
                        
                        doc = SimpleDocTemplate(pdf_filepath, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
                        elements = []
                        
                        elements.append(Paragraph("AR Balance Report", title_style))
                        elements.append(Spacer(1, 10))
                        
                        # [첫번째 블록 테이블]
                        b1_data = [
                            ["Nombre Cliente:", client_name, "Código Cliente:", client_code],
                            ["Fecha Límite (Max):", max_date_str, "Importe Total (EUR):", total_bal_fmt]
                        ]
                        b1_table = Table(b1_data, colWidths=[120, 250, 120, 100], hAlign='LEFT')
                        b1_table.setStyle(TableStyle([
                            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
                            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0,0), (-1,-1), 10),
                            ('TEXTCOLOR', (3,1), (3,1), colors.red if total_bal < 0 else colors.black),
                            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                        ]))
                        elements.append(b1_table)
                        elements.append(Spacer(1, 20))
                        
                        # [두번째 블록 테이블]
                        table_data = [available_pdf_cols]
                        custom_styles = [
                            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1557b0")),
                            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
                            ('FONTSIZE', (0,0), (-1,-1), 9),
                            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#DADCE0")),
                            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                            ('TOPPADDING', (0,0), (-1,-1), 6),
                        ]
                        
                        importe_idx = available_pdf_cols.index('Importe') if 'Importe' in available_pdf_cols else -1
                        balance_idx = available_pdf_cols.index('Balance') if 'Balance' in available_pdf_cols else -1
                        
                        if importe_idx != -1: custom_styles.append(('ALIGN', (importe_idx,1), (importe_idx,-1), 'RIGHT'))
                        if balance_idx != -1: custom_styles.append(('ALIGN', (balance_idx,1), (balance_idx,-1), 'RIGHT'))
                        
                        for r_idx, (_, row) in enumerate(group_df.iterrows(), start=1):
                            row_data = []
                            for c_idx, col in enumerate(available_pdf_cols):
                                val = row.get(col, '')
                                if col in ['Importe', 'Balance']:
                                    try:
                                        v = float(val)
                                        s = f"({abs(v):,.2f})" if v < 0 else f"{v:,.2f}"
                                        row_data.append(s)
                                        if v < 0:
                                            custom_styles.append(('TEXTCOLOR', (c_idx, r_idx), (c_idx, r_idx), colors.red))
                                    except:
                                        row_data.append(str(val))
                                else:
                                    row_data.append(str(val))
                            table_data.append(row_data)
                            
                        # 합계 라인 추가
                        total_row = ["Total"] + [""] * (len(available_pdf_cols) - 1)
                        if balance_idx != -1:
                            total_row[balance_idx] = f"({abs(total_bal):,.2f})" if total_bal < 0 else f"{total_bal:,.2f}"
                            
                        table_data.append(total_row)
                        last_row_idx = len(table_data) - 1
                        
                        custom_styles.extend([
                            ('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor("#F1F3F4")),
                            ('FONTNAME', (0, last_row_idx), (-1, last_row_idx), 'Helvetica-Bold'),
                            ('SPAN', (0, last_row_idx), (balance_idx - 1 if balance_idx > 0 else 0, last_row_idx)),
                            ('ALIGN', (0, last_row_idx), (balance_idx - 1 if balance_idx > 0 else 0, last_row_idx), 'RIGHT'),
                        ])
                        if total_bal < 0 and balance_idx != -1:
                            custom_styles.append(('TEXTCOLOR', (balance_idx, last_row_idx), (balance_idx, last_row_idx), colors.red))
                            
                        t = Table(table_data, repeatRows=1)
                        t.setStyle(TableStyle(custom_styles))
                        elements.append(t)
                        
                        doc.build(elements)
                        
                except ImportError:
                    self._log("❌ The 'reportlab' library is missing. PDF generation skipped.")

                self._set_status(85, self.t("saving"))
                self._log("Saving strictly formatted Excel file...")
                
                today_str = datetime.now().strftime('%Y%m%d')
                output_filepath = os.path.join(output_dir, f"{today_str} AR_Balance_Report.xlsx")

                with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
                    df_summary.to_excel(writer, sheet_name='Summary', index=False)
                    df.to_excel(writer, sheet_name='Report Format', index=False)
                    original_df.to_excel(writer, sheet_name='Original Data', index=False)
                    
                    workbook = writer.book
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    font_regular = Font(name='Aptos Narrow', size=10)
                    font_bold = Font(name='Aptos Narrow', size=10, bold=True)
                    font_red = Font(name='Aptos Narrow', size=10, color='FF0000')
                    
                    header_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
                    thin_side = Side(style='thin', color='DADCE0')
                    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                    
                    num_format = '#,##0.00;[Red](#,##0.00);"-"'

                    for sheet_name in ['Summary', 'Report Format']:
                        if sheet_name not in workbook.sheetnames:
                            continue
                        ws = workbook[sheet_name]
                        col_names = [str(cell.value).strip() for cell in ws[1]]
                        
                        max_col_letter = get_column_letter(ws.max_column)
                        ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"
                        
                        for cell in ws[1]:
                            cell.font = font_bold
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                        
                        for row in range(2, ws.max_row + 1):
                            for col in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row, column=col)
                                cell.font = font_regular
                                cell.border = thin_border
                                
                                col_name = col_names[col-1] if col-1 < len(col_names) else ''
                                
                                if col_name in ['Importe', 'Balance']:
                                    cell.number_format = num_format
                                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                                        cell.font = font_red
                                    cell.alignment = Alignment(horizontal='right')
                                elif col_name in ['Código Cliente', 'Factura', 'Tipología', 'Fecha emisión', 'Fecha Vencimiento']:
                                    cell.alignment = Alignment(horizontal='center')
                                else:
                                    cell.alignment = Alignment(horizontal='left')
                        
                        for col in ws.columns:
                            max_len = 0
                            for cell in col:
                                val_str = str(cell.value or '')
                                max_len = max(max_len, len(val_str))
                            col_letter = get_column_letter(col[0].column)
                            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

                self._set_status(100, self.t("completed_status"))
                self._log(f"[{datetime.now().strftime('%H:%M:%S')}] Task completed successfully.")
                messagebox.showinfo(self.t("done"), self.t("completed_msg").format(output_dir))

            except Exception as e:
                self._set_status(0, f"{self.t('error')}: {e}")
                self._log(f"❌ ERROR: {str(e)}")
                messagebox.showerror(self.t("error"), f"An error occurred:\n{str(e)}")
            finally:
                self.run_btn.config(state="normal", bg=PRIMARY)

        threading.Thread(target=worker, daemon=True).start()

if __name__ == "__main__":
    app = App()
    app.mainloop()
